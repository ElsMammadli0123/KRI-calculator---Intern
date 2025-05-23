"""
KRI CALCULATION SCRIPT 

"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from collections import defaultdict

# =====================
#  CONFIGURATION
# =====================
INPUT_FILE = "internship_task_data.xlsx"
OUTPUT_FILE = "KRI_FINAL_RESULTS.xlsx"
RTO_THRESHOLD = 120  # 2 hours in minutes
DATE_FORMAT = "%m/%d/%y %I:%M:%S %p"  # Updated to match AM/PM format in Excel
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True)

# =====================
#  DATA LOADING 
# =====================
def load_data():
    """Loads data with detailed error logging and collects all system names."""
    wb = load_workbook(INPUT_FILE)
    thresholds = {
        "MTBF": 90,
        "Monthly_Incidents": 3,
        "Exceed_RTO": 0,
        "Within_RTO": 3
    }
    incident_sheet = wb["incident_data"]
    incidents = []
    all_systems = set()
    current_system = None

    print("\n===== DATA LOADING LOG =====")
    for row_idx, row in enumerate(incident_sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if not any(row):
            continue

        # Track system names
        if row[0] and isinstance(row[0], str) and row[0].strip():
            current_system = row[0].strip()
            all_systems.add(current_system)
            print(f"Row {row_idx}: System set to '{current_system}'")
            continue

        # Parse incident data
        if row[1] and row[2] and row[3]:
            if current_system is None:
                print(f"Row {row_idx}: ERROR - No system name set for incident row: {row}")
                continue
            try:
                start = row[1] if isinstance(row[1], datetime) else datetime.strptime(row[1].strip(), DATE_FORMAT)
                end = row[2] if isinstance(row[2], datetime) else datetime.strptime(row[2].strip(), DATE_FORMAT)
                duration = int(row[3])

                incidents.append({
                    "system": current_system,
                    "start": start,
                    "end": end,
                    "duration": duration,
                    "month": f"{start.month:02d}-{start.year}"
                })
                print(f"Row {row_idx}: Loaded {current_system} incident ({duration} mins)")
            except Exception as e:
                print(f"Row {row_idx}: ERROR - {str(e)} | Row data: {row}")

    print(f"\nTOTAL INCIDENTS LOADED: {len(incidents)}")
    print(f"Systems found: {all_systems}")
    months = {i['month'] for i in incidents}
    print(f"Months found: {months}")
    return thresholds, incidents, all_systems

# =====================
#  KRI CALCULATIONS
# =====================
def calculate_kris(incidents, thresholds, all_systems):
    """Calculate all KRIs with threshold checks and ensure all systems are included."""

    # KRI1: MTBF
    mtbf = defaultdict(dict)
    for system in all_systems:
        sys_incidents = sorted([i for i in incidents if i["system"] == system], key=lambda x: x["start"])
        if len(sys_incidents) < 2:
            mtbf[system] = {"value": 0, "status": "N/A"}
        else:
            deltas = [(sys_incidents[i]["start"] - sys_incidents[i-1]["start"]).days for i in range(1, len(sys_incidents))]
            avg = sum(deltas)/len(deltas)
            status = "Breach" if avg < thresholds["MTBF"] else "OK"
            mtbf[system] = {"value": round(avg, 2), "status": status}

    # KRI2: Monthly Incidents
    monthly = defaultdict(list)
    for incident in incidents:
        key = (incident["system"], incident["month"])
        monthly[key].append(incident)

    monthly_results = {}
    for system in all_systems:
        months = {i["month"] for i in incidents if i["system"] == system}
        for month in months:
            count = len([i for i in incidents if i["system"] == system and i["month"] == month])
            status = "Breach" if count > thresholds["Monthly_Incidents"] else "OK"
            monthly_results[(system, month)] = {"count": count, "status": status}

    # KRI3 & KRI4: RTO Analysis
    rto_results = defaultdict(lambda: {"exceed": 0, "within": 0})
    for incident in incidents:
        system = incident["system"]
        if incident["duration"] > RTO_THRESHOLD:
            rto_results[system]["exceed"] += 1
        else:
            rto_results[system]["within"] += 1

    # Ensure all systems are present in rto_results
    for system in all_systems:
        if system not in rto_results:
            rto_results[system] = {"exceed": 0, "within": 0}
        rto_results[system]["exceed_status"] = "Breach" if rto_results[system]["exceed"] > 0 else "OK"
        rto_results[system]["within_status"] = "Breach" if rto_results[system]["within"] > 3 else "OK"

    print("\n===== MTBF Results =====")
    for system, data in mtbf.items():
        print(f"{system}: {data}")

    print("\n===== Monthly Incident Results =====")
    for key, data in monthly_results.items():
        print(f"{key}: {data}")

    print("\n===== RTO Results =====")
    for system, data in rto_results.items():
        print(f"{system}: {data}")

    return mtbf, monthly_results, rto_results

# =====================
#  EXCEL REPORT GENERATION
# =====================
def create_excel_report(mtbf, monthly, rto, thresholds):
    """Generates formatted Excel output with guaranteed visibility"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: MTBF Analysis
    ws1 = wb.create_sheet("KRI1 - MTBF")
    headers = ["System", "MTBF (Days)", "Threshold", "Status"]
    ws1.append(headers)
    for system, data in mtbf.items():
        ws1.append([system, data["value"], thresholds["MTBF"], data["status"]])
        if data["status"] == "Breach":
            ws1.cell(row=ws1.max_row, column=4).fill = RED_FILL
    
    # Sheet 2: Monthly Incidents
    ws2 = wb.create_sheet("KRI2 - Monthly")
    headers = ["System", "Incident Count", "Threshold", "Status"]
    ws2.append(headers)
    for system in mtbf.keys():
        total_count = sum(data["count"] for (sys, _), data in monthly.items() if sys == system)
        status = "Breach" if total_count > thresholds["Monthly_Incidents"] else "OK"
        ws2.append([system, total_count, thresholds["Monthly_Incidents"], status])
        if status == "Breach":
            ws2.cell(row=ws2.max_row, column=4).fill = RED_FILL
    
    # Sheet 3: RTO Analysis
    ws3 = wb.create_sheet("KRI3-KRI4 - RTO")
    headers = ["System", "Exceeded RTO", "Within RTO", "Threshold (Exceed)", "Threshold (Within)", "Exceed Status", "Within Status"]
    ws3.append(headers)
    for system, data in rto.items():
        ws3.append([
            system,
            data["exceed"],
            data["within"],
            thresholds["Exceed_RTO"],
            thresholds["Within_RTO"],
            data["exceed_status"],
            data["within_status"]
        ])
        if data["exceed_status"] == "Breach":
            ws3.cell(row=ws3.max_row, column=6).fill = RED_FILL
        if data["within_status"] == "Breach":
            ws3.cell(row=ws3.max_row, column=7).fill = RED_FILL
    
    # Formatting for all sheets
    for sheet in wb.worksheets:
        # Set column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        # Bold headers
        for cell in sheet[1]:
            cell.font = HEADER_FONT
    
    wb.save(OUTPUT_FILE)
    print(f"\nReport saved to: {OUTPUT_FILE}")

# =====================
#  MAIN EXECUTION
# =====================
if __name__ == "__main__":
    try:
        thresholds, incidents, all_systems = load_data()
        mtbf, monthly, rto = calculate_kris(incidents, thresholds, all_systems)
        create_excel_report(mtbf, monthly, rto, thresholds)
    except Exception as e:
        print(f"FATAL ERROR: {e}")
